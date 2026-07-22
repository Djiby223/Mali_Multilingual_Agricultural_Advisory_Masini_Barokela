from openpyxl import load_workbook
from pathlib import Path
import zipfile

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILE = PROJECT_ROOT / "data" / "Masini_Barokela_Master_Knowledge_Base.xlsx"

print("File:", EXCEL_FILE)
print("Exists:", EXCEL_FILE.exists())
print("ZIP file:", zipfile.is_zipfile(EXCEL_FILE))

wb = load_workbook(EXCEL_FILE)

print("SUCCESS!")
print("Sheets:", wb.sheetnames)